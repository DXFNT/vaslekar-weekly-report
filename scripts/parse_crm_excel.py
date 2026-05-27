#!/usr/bin/env python3
"""Parse Váš Lekár CRM Excel exports into JSON aggregates for the weekly report.

Inputs:
    orders.xlsx     — "zoznam objednavok klientov" — one row per order line
                       (Dátum úhrady, Meno, Priezvisko, Adresa, Tel., Email, Suma, Produkt, Súhlas)
    products.xlsx   — "export predanych produktov" — per-day product matrix (not currently used,
                       kept for future per-service breakdowns when format stabilizes)

Outputs JSON:
{
    "closed_week": {
        "start": "2026-05-18",
        "end":   "2026-05-24",
        "orders": 54,
        "revenue": 6770.05,
        "aov": 125.37,
        "products": [{"name": "...", "count": 29, "revenue": 1711.0}, ...],
        "jednorazove_by_spec": {"ORL": 11, "Angiológia": 7, ...}
    },
    "prev_week": { ... same shape ... },
    "wow": {
        "orders_pct": -25.0,
        "revenue_pct": -19.0,
        "aov_pct": 8.0
    }
}

Usage:
    python parse_crm_excel.py --orders orders.xlsx --products products.xlsx --week-ending 2026-05-24
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


def parse_dt(value) -> date | None:
    """Robust parsing of the 'Dátum úhrady' column (datetime or 'DD.MM.YYYY HH:MM:SS' string)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    s = str(value)
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y"):
        try:
            return datetime.strptime(s[: len(fmt) + 4], fmt).date()
        except ValueError:
            continue
    return None


def parse_amount(value) -> float:
    """Parses '59,00' / 59.0 / '-44,90' → float."""
    if isinstance(value, (int, float)):
        return float(value)
    if value is None:
        return 0.0
    s = str(value).replace(",", ".").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def product_base(product: str) -> str:
    """'1x jednorazový vstup - ORL' → '1x jednorazový vstup'."""
    if not product:
        return "Unknown"
    return product.split(" - ", 1)[0].strip()


def specialization(product: str) -> str | None:
    """Extract specialization from 'Nx jednorazový vstup - <Spec>'. Returns None if not a jednorazový vstup."""
    if not product or "jednorazový vstup" not in product.lower():
        return None
    if " - " in product:
        return product.split(" - ", 1)[1].strip()
    return "Bez špecializácie"


def aggregate(rows, start: date, end: date) -> dict:
    """Aggregate orders for [start, end] inclusive."""
    seen_payments = set()  # dedupe orders that have multiple rows (e.g. one row for product, one for promo discount)
    orders = 0
    revenue = 0.0
    products: dict[str, dict] = defaultdict(lambda: {"count": 0, "revenue": 0.0})
    jednorazove: dict[str, int] = defaultdict(int)

    for r in rows:
        d, dt_str, email, amount, product = r["date"], r["dt_str"], r["email"], r["amount"], r["product"]
        if not d or d < start or d > end:
            continue
        is_promo = product and "Promo" in product
        payment_key = (dt_str, email)
        # One order per unique payment, but include all rows (including promo discounts) in revenue
        if payment_key not in seen_payments and not is_promo:
            orders += 1
            seen_payments.add(payment_key)
        revenue += amount
        if not is_promo and product:
            base = product_base(product)
            products[base]["count"] += 1
            products[base]["revenue"] += amount
            spec = specialization(product)
            if spec:
                jednorazove[spec] += 1

    return {
        "start": start.isoformat(),
        "end": end.isoformat(),
        "orders": orders,
        "revenue": round(revenue, 2),
        "aov": round(revenue / orders, 2) if orders else 0.0,
        "products": sorted(
            (
                {"name": k, "count": v["count"], "revenue": round(v["revenue"], 2)}
                for k, v in products.items()
            ),
            key=lambda p: -p["revenue"],
        ),
        "jednorazove_by_spec": dict(sorted(jednorazove.items(), key=lambda x: -x[1])),
    }


def load_rows(orders_xlsx: Path) -> list[dict]:
    wb = load_workbook(orders_xlsx, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 8:
                continue
            dt_str, _, _, _, _, email, suma, produkt = row[:8]
            d = parse_dt(dt_str)
            rows.append(
                {
                    "date": d,
                    "dt_str": dt_str,
                    "email": email,
                    "amount": parse_amount(suma),
                    "product": str(produkt) if produkt else "",
                }
            )
    return rows


def pct_change(new: float, old: float) -> float | None:
    if old == 0:
        return None
    return round((new - old) / old * 100, 1)


def main() -> int:
    p = argparse.ArgumentParser(description="Parse Váš Lekár CRM Excel → JSON for weekly report")
    p.add_argument("--orders", required=True, type=Path, help="Path to zoznam_objednavok_klientov.xlsx")
    p.add_argument("--products", type=Path, help="Path to export_predanych_produktov.xlsx (optional, reserved)")
    p.add_argument("--week-ending", required=True, help="Last day of the closed week (YYYY-MM-DD, must be a Sunday)")
    p.add_argument("--out", type=Path, default=Path("crm_aggregates.json"))
    args = p.parse_args()

    closed_end = date.fromisoformat(args.week_ending)
    if closed_end.weekday() != 6:
        print(f"WARN: --week-ending {closed_end} is not a Sunday (got weekday={closed_end.weekday()})", file=sys.stderr)
    closed_start = closed_end - timedelta(days=6)
    prev_end = closed_start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=6)

    rows = load_rows(args.orders)
    closed = aggregate(rows, closed_start, closed_end)
    prev = aggregate(rows, prev_start, prev_end)

    result = {
        "closed_week": closed,
        "prev_week": prev,
        "wow": {
            "orders_pct": pct_change(closed["orders"], prev["orders"]),
            "revenue_pct": pct_change(closed["revenue"], prev["revenue"]),
            "aov_pct": pct_change(closed["aov"], prev["aov"]),
        },
    }
    args.out.write_text(json.dumps(result, ensure_ascii=False, indent=2))
    print(f"Wrote {args.out}")
    print(f"  Closed week ({closed_start}…{closed_end}): {closed['orders']} orders, €{closed['revenue']}")
    print(f"  Prev week    ({prev_start}…{prev_end}): {prev['orders']} orders, €{prev['revenue']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
